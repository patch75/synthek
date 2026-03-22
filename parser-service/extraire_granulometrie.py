"""
Synthek — extraire_granulometrie.py
Pipeline import granulométrie : fichier architecte (Excel) → JSON normalisé contrat D1
via extraction texte brut (Python) + extraction sémantique (Sonnet 4.6)

Workflow :
  Appel 1 (regroupement_valide=None, nom_feuille=None|"X") :
    a) Si plusieurs feuilles éligibles et nom_feuille absent :
         → { etape: "selection_feuille", feuilles_disponibles, feuille_suggeree }
    b) Sinon :
         → Extrait texte brut + appelle Sonnet
         → { etape: "validation", batiments: [...], total_logements, ... }

  Appel 2 — nouveau format (regroupement_valide = list de batiment objects) :
    → Valide les données confirmées par l'utilisateur → retourne JSON D1 final

  Appel 2 — rétrocompat (regroupement_valide = dict {groupe: [montees]}) :
    → Re-extrait + appelle Sonnet → retourne JSON D1 final
"""

import io
import re
import os
import json
import logging
import datetime
from typing import Optional

import openpyxl
import anthropic

logger = logging.getLogger(__name__)

# ─────────────────────────────────────────────────────────────────
# CONSTANTES
# ─────────────────────────────────────────────────────────────────

PREFIXES_FEUILLES_EXCLUES = (
    'calcul', 'détail', 'detail', 'graphique', 'récap', 'recap', 'total'
)
LIMITE_TEXTE_BRUT = 25_000
MODELE_SONNET = 'claude-sonnet-4-6'

PROMPT_SYSTEME = """Ne réfléchis pas à voix haute. Commence ta réponse immédiatement par { sans aucun texte avant.

Tu es un parser de donnees immobilieres specialise en programmes de logements.
Tu recois le contenu brut d'un fichier Excel architecte au format positionnel groupé par ligne :
  L01 | col00=BATIMENT | col02=A1 | col05=A2 | col14=B | col23=C (LLS)
  L02 | col00=NIVEAU | col02=RDC | col03=R+1 | col05=RDC | col08=R+1 | col11=R+2
  L03 | col00=N° | col02=001 | col03=101 (BRS) | col05=001 | col06=002 | col07=003
Chaque ligne = une ligne Excel non vide. col00 = colonne A, col01 = B, etc. (index 0-based).
Les gaps entre indices de colonnes indiquent les cellules vides (ex : col02 à col05 = A1 occupe 3 colonnes).

Extrais les informations suivantes par batiment ou groupe de batiments :
- nom du batiment ou groupe (ex: A, BAT A, Batiment A)
- montees : sous-entrees du batiment si presentes (ex: A1, A2) — [] si absent ou SANS OBJET
- nos_comptes : liste exhaustive et exacte de tous les N° de logements que tu as lus pour ce batiment
- nb_logements : doit etre exactement egal a len(nos_comptes)
- LLI, LLS, BRS, acces_std, acces_premium, villas : deduits des annotations dans nos_comptes

Regles strictes :
- Lecture colonnes : la ligne BATIMENT definit les groupes — chaque nom (A1, A2, B...) occupe les colonnes depuis son index jusqu'au prochain nom. Les N° sous ces colonnes appartiennent a ce batiment.
- nos_comptes : liste TOUS les N° du fichier pour ce batiment — ne rien inventer, ne rien omettre, ne pas dupliquer entre batiments
- Financements deduits de nos_comptes : "(LLS)" → LLS, "(BRS)" → BRS, "(LLI)" → LLI, sans annotation → accession libre (acces_std ou acces_premium), entier 1-2 chiffres → villa
- Nom batiment ou montee contient "(LLS)"/"(LLI)"/"(BRS)" → tous logements sans annotation de ce batiment = ce financement
- PREMIUM = logement au dernier niveau de sa montee sans annotation sociale
- VILLA = maison individuelle — Section VILLAS : entiers simples sur la ligne suivant "VILLAS" = numeros de villas
- SANS OBJET dans montees → montees: []
- Ignorer lignes TOTAL et recapitulatives
- Si une donnee est absente ou non deductible avec certitude → null
- Fiabilite "haute" si nos_comptes rempli, "estimee" si compte depuis typologies, "incomplete" si financements tous null
- Retourne UNIQUEMENT le JSON valide, sans texte avant ni apres, sans backticks

Format de sortie exact :
{
  "projet": null,
  "source": "nom_fichier",
  "batiments": [
    {
      "nom": "A",
      "montees": ["A1", "A2"],
      "nos_comptes": ["001", "002 (BRS)", "101", "102 (LLS)"],
      "nb_logements": 4,
      "LLI": null,
      "LLS": 1,
      "BRS": 1,
      "acces_std": 2,
      "acces_premium": null,
      "villas": 0,
      "fiabilite": "haute",
      "section_cctp": null,
      "feuilles_dpgf": [],
      "systeme_chauffage": null,
      "systeme_vmc": null,
      "regulation": null,
      "notes": null
    }
  ],
  "total_logements": 4,
  "donnees_manquantes": [],
  "hypotheses": []
}"""


# ─────────────────────────────────────────────────────────────────
# ÉTAPE 1 — EXTRACTION TEXTE BRUT EXCEL
# ─────────────────────────────────────────────────────────────────

def _feuille_eligible(nom: str) -> bool:
    n = nom.lower().strip()
    return not any(n.startswith(p) for p in PREFIXES_FEUILLES_EXCLUES)


def _extraire_texte_brut_excel(file_bytes: bytes, nom_feuille: str = None) -> dict:
    """
    Lit le fichier Excel et produit un texte positionnel cellule par cellule.

    Retourne l'un de ces deux formats :
      {'texte': str, 'tronque': bool, 'feuille': str}
          — si une feuille cible est identifiée et extraite
      {'feuilles_disponibles': [...], 'feuille_suggeree': str}
          — si plusieurs feuilles éligibles et nom_feuille non fourni
    """
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    feuilles_eligibles = [s for s in wb.sheetnames if _feuille_eligible(s)]

    # Résoudre la feuille cible
    if nom_feuille:
        cible = nom_feuille
    elif len(feuilles_eligibles) == 1:
        cible = feuilles_eligibles[0]
    elif not feuilles_eligibles:
        cible = wb.sheetnames[0]  # fallback : aucune feuille exclue
    else:
        # Plusieurs feuilles éligibles → retourner la liste pour sélection utilisateur
        suggeree = next(
            (s for s in feuilles_eligibles if 'surfaces pro' in s.lower()),
            feuilles_eligibles[0],
        )
        return {
            'feuilles_disponibles': feuilles_eligibles,
            'feuille_suggeree': suggeree,
        }

    if cible not in wb.sheetnames:
        raise ValueError(f"Feuille '{cible}' introuvable. Disponibles : {list(wb.sheetnames)}")

    ws = wb[cible]
    lignes_dump = []
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        cellules = []
        for j, val in enumerate(row):  # j = index 0-based
            if val is None:
                continue
            if isinstance(val, (datetime.date, datetime.datetime)):
                continue
            if isinstance(val, float):
                continue
            s = str(val).strip()
            if not s or s.startswith('='):
                continue
            cellules.append(f'col{j:02d}={s}')
        if cellules:
            lignes_dump.append(f'L{i:02d} | ' + ' | '.join(cellules))

    texte = '\n'.join(lignes_dump)
    tronque = len(texte) > LIMITE_TEXTE_BRUT
    if tronque:
        texte = texte[:LIMITE_TEXTE_BRUT]

    return {'texte': texte, 'tronque': tronque, 'feuille': cible}


# ─────────────────────────────────────────────────────────────────
# ÉTAPE 2 — EXTRACTION SÉMANTIQUE VIA SONNET
# ─────────────────────────────────────────────────────────────────

def _compter_logements_individuels(dump: str) -> tuple:
    """
    Compte les N° de logements individuels depuis le dump positionnel groupé.

    Format attendu : "L01 | col00=BATIMENT | col02=A1 | col05=A2"
    col00 = colonne A (index 0-based).

    Détection lignes N° :
      - Primaire : col00 contient "N°" ou variante (N°, No, N.)
      - Fallback : ligne avec ≥3 tokens col\d+ de type N° 3-chiffres hors col00

    Retourne (count, tous_nos) :
      count    : int — total N° détectés (0 si format typologies/comptage)
      tous_nos : list[str] — liste brute exhaustive pour contrainte Sonnet
    """
    lignes_par_ln = {}   # {ln: [(col_idx, val)]}
    lignes_villas = set()
    lignes_numero = set()

    for line in dump.split('\n'):
        line = line.strip()
        if not line.startswith('L'):
            continue
        parts = line.split(' | ')
        if not parts:
            continue
        try:
            ln = int(parts[0][1:])
        except ValueError:
            continue
        cellules = []
        for part in parts[1:]:
            idx_str, _, val = part.partition('=')
            if not idx_str.startswith('col'):
                continue
            try:
                col_idx = int(idx_str[3:])
            except ValueError:
                continue
            val = val.strip()
            if not val:
                continue
            cellules.append((col_idx, val))
            if val.upper() == 'VILLAS':
                lignes_villas.add(ln)
            if col_idx == 0 and re.match(r'^N.{0,2}$', val):
                lignes_numero.add(ln)
        lignes_par_ln[ln] = cellules

    pat_collectif = re.compile(r'^\d{3,}(\s*\([^)]+\))?$')
    pat_villa_num = re.compile(r'^\d{1,2}$')

    # Fallback : ligne avec ≥3 N° 3-chiffres hors col00
    for ln, cellules in lignes_par_ln.items():
        if ln in lignes_numero or ln in lignes_villas:
            continue
        nb_nos = sum(1 for col_idx, val in cellules if col_idx > 0 and pat_collectif.match(val))
        if nb_nos >= 3:
            lignes_numero.add(ln)

    tous_nos = []
    for ln, cellules in lignes_par_ln.items():
        is_numero_row = ln in lignes_numero
        is_villa_row  = (ln - 1) in lignes_villas or (ln - 2) in lignes_villas
        for col_idx, val in cellules:
            if col_idx == 0:
                continue
            if is_numero_row and pat_collectif.match(val):
                tous_nos.append(val)
            elif is_villa_row and pat_villa_num.match(val):
                tous_nos.append(val)

    return len(tous_nos), tous_nos


def _formater_contrainte_nos(nb: int, tous_nos: list) -> str:
    """Formate la contrainte N° exhaustive pour le prompt Sonnet."""
    if not tous_nos:
        return ''
    return (
        f'\n\nContrainte stricte — {nb} N° de logements détectés par Python (liste exhaustive) :\n'
        f'{tous_nos}\n'
        'Règles absolues :\n'
        '- nos_comptes de chaque bâtiment = les N°s de cette liste qui lui appartiennent\n'
        '- Chaque N° doit figurer dans nos_comptes d\'exactement UN bâtiment (pas de doublon)\n'
        '- nb_logements = len(nos_comptes) pour chaque bâtiment\n'
        f'- total_logements DOIT être exactement {nb}'
    )


def _verifier_et_corriger_batiments(batiments: list, nb_python: int) -> tuple:
    """
    Vérifie la cohérence des nos_comptes et dérive les financements.

    Règle 1 — nb_logements = len(nos_comptes) : corrige silencieusement
    Financement global — nom ou montées contient "(LLS)"/"(LLI)"/"(BRS)" :
      tous les libres de ce bâtiment = ce financement
    Warning [NOS-PREM] — std+premium ≠ libres : tout mis en acces_std

    Dérive financements depuis annotations nos_comptes :
      "(LLS)" → LLS, "(BRS)" → BRS, "(LLI)" → LLI
      entier 1-2 chiffres → villas, sans annotation → accession libre

    Retourne (batiments_corrigés, warnings).
    """
    warnings = []
    pat_villa_num = re.compile(r'^\d{1,2}$')

    # Règle 1 : nb_logements = len(nos_comptes)
    for b in batiments:
        nos = b.get('nos_comptes') or []
        nb  = b.get('nb_logements')
        if nos and nb != len(nos):
            warnings.append(
                f"[NOS-R1] {b.get('nom', '?')} : nb_logements={nb} "
                f"≠ len(nos_comptes)={len(nos)} → corrigé"
            )
            b['nb_logements'] = len(nos)

    # R2 et R3 supprimées — les N° 001/101/201 se répètent légitimement dans chaque bâtiment
    # (chaque montée a son propre RDC/R+1/R+2), ce ne sont pas des doublons inter-bâtiments.

    # Vérification et dérivation financements
    # Règle : si somme(financements Sonnet) == nb_logements → Sonnet est cohérent → garder
    # Sinon → dériver depuis annotations dans nos_comptes (fallback)
    for b in batiments:
        nos = b.get('nos_comptes') or []
        nb  = b.get('nb_logements') or 0

        somme_sonnet = (
            (b.get('LLI') or 0) + (b.get('LLS') or 0) + (b.get('BRS') or 0) +
            (b.get('acces_std') or 0) + (b.get('acces_premium') or 0) +
            (b.get('villas') or 0)
        )

        if somme_sonnet == nb:
            # Sonnet cohérent → conserver sa répartition, normaliser None → 0
            b['LLI']           = b.get('LLI') or 0
            b['LLS']           = b.get('LLS') or 0
            b['BRS']           = b.get('BRS') or 0
            b['acces_std']     = b.get('acces_std') or 0
            b['acces_premium'] = b.get('acces_premium') or 0
            b['villas']        = b.get('villas') or 0
        else:
            # Sonnet incohérent → dériver depuis annotations nos_comptes
            if not nos:
                continue
            lls    = sum(1 for n in nos if re.search(r'\(LLS\)', n, re.IGNORECASE))
            brs    = sum(1 for n in nos if re.search(r'\(BRS\)', n, re.IGNORECASE))
            lli    = sum(1 for n in nos if re.search(r'\(LLI\)', n, re.IGNORECASE))
            villas = sum(1 for n in nos if pat_villa_num.match(n))
            libres = sum(1 for n in nos if not re.search(r'\(', n) and not pat_villa_num.match(n))

            # Financement global depuis nom bâtiment ou montées
            sources_nom = [b.get('nom', '')] + list(b.get('montees') or [])
            financement_global = None
            for source in sources_nom:
                if re.search(r'\(LLS\)', source, re.IGNORECASE):
                    financement_global = 'LLS'; break
                if re.search(r'\(LLI\)', source, re.IGNORECASE):
                    financement_global = 'LLI'; break
                if re.search(r'\(BRS\)', source, re.IGNORECASE):
                    financement_global = 'BRS'; break
            if financement_global and libres > 0:
                if financement_global == 'LLS':   lls += libres
                elif financement_global == 'LLI': lli += libres
                elif financement_global == 'BRS': brs += libres
                libres = 0

            b['LLS']           = lls or 0
            b['BRS']           = brs or 0
            b['LLI']           = lli or 0
            b['acces_std']     = libres or 0
            b['acces_premium'] = 0
            b['villas']        = villas or (b.get('villas') or 0)
            warnings.append(
                f"[FIN-DERIVE] {b.get('nom','?')} : "
                f"somme Sonnet ({somme_sonnet}) ≠ nb_logements ({nb}) → financements recalculés"
            )

        # Fiabilité haute si nb_logements connu
        if b.get('nb_logements') is not None:
            b['fiabilite'] = 'haute'

    return batiments, warnings


def _appeler_sonnet(texte_brut: str, nom_fichier: str, nb_logements_detectes: int = 0, tous_nos: list = None) -> dict:
    """
    Envoie le texte positionnel à Sonnet 4.6 et retourne le dict Python parsé.
    Si nb_logements_detectes > 0, ajoute la liste exhaustive des N°s comme contrainte.
    """
    contrainte = ''
    if nb_logements_detectes > 0:
        contrainte = _formater_contrainte_nos(nb_logements_detectes, tous_nos or [])

    client = anthropic.Anthropic(api_key=os.environ.get('ANTHROPIC_API_KEY'))
    message = client.messages.create(
        model=MODELE_SONNET,
        max_tokens=4000,
        system=PROMPT_SYSTEME,
        messages=[{
            'role': 'user',
            'content': f'Fichier : {nom_fichier}{contrainte}\n\n{texte_brut}',
        }],
    )

    raw = message.content[0].text.strip()
    # Nettoyer les backticks markdown éventuels
    raw = re.sub(r'^```(?:json)?\s*', '', raw)
    raw = re.sub(r'\s*```$', '', raw)

    # Parsing direct
    try:
        return json.loads(raw)
    except json.JSONDecodeError:
        pass

    # Fallback : extraire le bloc JSON depuis le premier { jusqu'à la fin
    idx = raw.find('{')
    if idx > 0:
        try:
            return json.loads(raw[idx:])
        except json.JSONDecodeError:
            pass

    raise ValueError(
        f"Sonnet n'a pas retourné un JSON valide.\n"
        f"Réponse brute (500 premiers chars) : {raw[:500]}"
    )


# ─────────────────────────────────────────────────────────────────
# ÉTAPE 3 — VALIDATION
# ─────────────────────────────────────────────────────────────────

def _valider(batiments: list) -> list:
    """
    V1 : nb_logements = somme typologies — uniquement si fiabilite == "haute" ET financements non null.
    V5 : aucun financement identifié — skip si fiabilite "incomplete" ou "estimee".
    """
    warnings = []
    for b in batiments:
        fiabilite = b.get('fiabilite', 'incomplete')
        nb      = b.get('nb_logements') or 0
        LLI     = b.get('LLI') or 0
        LLS     = b.get('LLS') or 0
        BRS     = b.get('BRS') or 0
        std     = b.get('acces_std') or 0
        premium = b.get('acces_premium') or 0
        villas  = b.get('villas') or 0

        financements_null = all(
            b.get(k) is None
            for k in ('LLI', 'LLS', 'BRS', 'acces_std', 'acces_premium', 'villas')
        )

        if fiabilite == 'haute' and not financements_null:
            somme = LLI + LLS + BRS + std + premium + villas
            if somme != nb:
                warnings.append(
                    f"V1 [{b.get('nom', '?')}] : nb_logements={nb} ≠ somme typologies={somme}"
                )

        if fiabilite not in ('incomplete', 'estimee') and not financements_null:
            somme = LLI + LLS + BRS + std + premium + villas
            if nb > 0 and somme == 0:
                warnings.append(f"V5 [{b.get('nom', '?')}] : aucun financement identifié")

    return warnings


# ─────────────────────────────────────────────────────────────────
# STUB — conservé pour compatibilité import main.py
# ─────────────────────────────────────────────────────────────────

def proposer_regroupement(montees: list) -> dict:
    """Conservé pour compatibilité avec main.py — non utilisé dans le pipeline LLM."""
    return {}


# ─────────────────────────────────────────────────────────────────
# HELPERS INTERNES
# ─────────────────────────────────────────────────────────────────

def _run_extraction_sonnet(
    file_bytes: bytes,
    nom_fichier: str,
    nom_feuille: str = None,
    force_feuille_suggeree: bool = False,
) -> dict:
    """
    Extrait le texte brut et appelle Sonnet.
    Retourne soit un dict avec 'feuilles_disponibles' (sélection requise),
    soit un dict avec 'batiments', 'total_logements', etc.
    """
    resultat_extraction = _extraire_texte_brut_excel(file_bytes, nom_feuille)

    # Multi-feuilles sans cible forcée
    if 'feuilles_disponibles' in resultat_extraction:
        if force_feuille_suggeree:
            # En mode rétrocompat Appel 2 : utiliser la feuille suggérée automatiquement
            resultat_extraction = _extraire_texte_brut_excel(
                file_bytes, resultat_extraction['feuille_suggeree']
            )
        else:
            return resultat_extraction  # renvoi liste feuilles pour sélection UI

    texte_brut = resultat_extraction['texte']
    tronque = resultat_extraction.get('tronque', False)
    feuille_utilisee = resultat_extraction.get('feuille', nom_feuille or '')

    nb_individuels, tous_nos = _compter_logements_individuels(texte_brut)
    json_sonnet = _appeler_sonnet(texte_brut, nom_fichier, nb_logements_detectes=nb_individuels, tous_nos=tous_nos)

    print("=== JSON SONNET BRUT ===", flush=True)
    print(json.dumps(json_sonnet, ensure_ascii=False, indent=2), flush=True)
    print("=== FIN JSON SONNET ===", flush=True)

    batiments = json_sonnet.get('batiments', [])

    # Vérification et correction via nos_comptes
    if nb_individuels > 0:
        batiments, warnings_nos = _verifier_et_corriger_batiments(batiments, nb_individuels)
    else:
        warnings_nos = []

    # Total : Python fait foi si N° détectés, sinon Sonnet
    if nb_individuels > 0:
        total = nb_individuels
    else:
        total = json_sonnet.get('total_logements') or sum(
            (b.get('nb_logements') or 0) for b in batiments
        )

    donnees_manquantes = list(json_sonnet.get('donnees_manquantes', []))
    donnees_manquantes.extend(warnings_nos)
    hypotheses = list(json_sonnet.get('hypotheses', []))

    if tronque:
        donnees_manquantes.append(
            f"Fichier tronqué à {LIMITE_TEXTE_BRUT} caractères — données potentiellement incomplètes"
        )

    return {
        'projet': json_sonnet.get('projet'),
        'source': nom_fichier,
        'feuille': feuille_utilisee,
        'batiments': batiments,
        'total_logements': total,
        'donnees_manquantes': donnees_manquantes,
        'hypotheses': hypotheses,
    }


# ─────────────────────────────────────────────────────────────────
# POINT D'ENTRÉE PRINCIPAL
# ─────────────────────────────────────────────────────────────────

def extraire_granulometrie(
    file_bytes: bytes,
    nom_fichier: str,
    regroupement_valide=None,
    nom_feuille: str = None,
) -> dict:
    """
    Pipeline granulométrie LLM universel.

    Appel 1 (regroupement_valide=None) :
      → Extrait texte brut + appelle Sonnet
      → Retourne {etape: "selection_feuille"} si multi-feuilles
      → Retourne {etape: "validation", batiments: [...]} sinon

    Appel 2 nouveau format (regroupement_valide = list) :
      → Valide les batiments confirmés/édités par l'utilisateur
      → Retourne JSON D1 final (pas de second appel Sonnet)

    Appel 2 rétrocompat (regroupement_valide = dict {groupe: [montees]}) :
      → Re-extrait + appelle Sonnet → retourne JSON D1 final
    """
    ext = nom_fichier.lower().rsplit('.', 1)[-1] if '.' in nom_fichier else ''
    if ext == 'pdf':
        raise NotImplementedError("Parsing PDF non implémenté. Utiliser un fichier Excel.")
    if ext not in ('xlsx', 'xlsm', 'xls'):
        raise ValueError(f"Format non supporté : {ext}. Acceptés : xlsx, xlsm, xls")

    # ── Appel 2 nouveau format : l'utilisateur renvoie la liste de batiments validée ──
    if isinstance(regroupement_valide, list):
        batiments = regroupement_valide
        warnings = _valider(batiments)
        total = sum((b.get('nb_logements') or 0) for b in batiments)
        return {
            'projet': None,
            'source': nom_fichier,
            'batiments': batiments,
            'total_logements': total,
            'donnees_manquantes': warnings,
            'hypotheses': [
                'Données extraites via Sonnet 4.6 — vérifiées et confirmées par le BET',
            ],
        }

    # ── Appel 1 + Appel 2 rétrocompat (dict) : extraction texte + Sonnet ──
    force_suggeree = isinstance(regroupement_valide, dict)
    resultat = _run_extraction_sonnet(
        file_bytes, nom_fichier, nom_feuille, force_feuille_suggeree=force_suggeree
    )

    # Cas : sélection de feuille nécessaire (Appel 1 multi-feuilles uniquement)
    if 'feuilles_disponibles' in resultat:
        return {
            'etape': 'selection_feuille',
            'feuilles_disponibles': resultat['feuilles_disponibles'],
            'feuille_suggeree': resultat['feuille_suggeree'],
            'message': 'Plusieurs feuilles disponibles. Confirmer la feuille de référence.',
        }

    batiments          = resultat['batiments']
    total              = resultat['total_logements']
    donnees_manquantes = resultat['donnees_manquantes']
    hypotheses         = resultat['hypotheses']

    # ── Appel 1 : retourner pour validation utilisateur ──
    if regroupement_valide is None:
        return {
            'etape': 'validation',
            'projet': resultat.get('projet'),
            'source': nom_fichier,
            'feuille': resultat.get('feuille', ''),
            'batiments': batiments,
            'total_logements': total,
            'donnees_manquantes': donnees_manquantes,
            'hypotheses': hypotheses,
        }

    # ── Appel 2 rétrocompat (dict) : retourner JSON D1 final avec données Sonnet ──
    warnings = _valider(batiments)
    warnings.extend(donnees_manquantes)
    return {
        'projet': resultat.get('projet'),
        'source': nom_fichier,
        'batiments': batiments,
        'total_logements': total,
        'donnees_manquantes': warnings,
        'hypotheses': hypotheses,
    }
