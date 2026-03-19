"""
MOE.AI — Module comparaison CCTP ↔ DPGF pour Synthek
=====================================================
Fichier : parser-service/comparaison_cctp_dpgf.py

Version Synthek — pas de CLI, pas de rapport Excel.
Accepte des bytes (base64 décodé) en entrée.
Extrait de moteur_comparaison_cctp_dpgf.py V2.1
"""

import io
import re
from dataclasses import dataclass, field
from typing import Optional

import openpyxl
from docx import Document

from equivalences_fluides import (
    normaliser,
    sont_equivalents,
    sont_non_equivalents,
    est_ligne_exclue,
    comparer_marques,
    forme_canonique_souple,
    comparer_puissances,
    extraire_puissance,
    extraire_diametre,
    est_designation_incertaine,
)


# ============================================================
# STRUCTURES DE DONNÉES
# ============================================================

@dataclass
class AssertionCCTP:
    section: str
    titre_article: str
    texte: str
    batiment: str
    type_canonique: Optional[str] = None
    puissance_kw: Optional[float] = None
    diametre_dn: Optional[int] = None


@dataclass
class LigneDPGF:
    designation: str
    feuille: str
    ligne: int
    batiment: str
    type_canonique: Optional[str] = None
    puissance_kw: Optional[float] = None
    diametre_dn: Optional[int] = None


@dataclass
class Alerte:
    code: str
    criticite: str
    batiment: str
    cctp_section: str = ""
    cctp_texte: str = ""
    dpgf_feuille: str = ""
    dpgf_ligne: int = 0
    dpgf_texte: str = ""
    motif: str = ""
    regle: str = ""
    methode: str = "Python"
    confiance: int = 0

    @property
    def signature(self) -> str:
        return f"{self.code}|{self.batiment}|{self.cctp_texte[:80]}|{self.dpgf_texte[:80]}"


@dataclass
class ProgrammeBatiment:
    nom: str
    section_cctp: str
    feuilles_dpgf: list
    nb_logements_total: int = 0
    types_logements: dict = field(default_factory=dict)
    systeme_chauffage: str = ""
    notes: str = ""


# ============================================================
# EXTRACTION PROGRAMME
# ============================================================

def extraire_programme(config: dict) -> list[ProgrammeBatiment]:
    if "programme" not in config:
        return []
    return [
        ProgrammeBatiment(
            nom=b.get("nom", ""),
            section_cctp=b.get("section_cctp", ""),
            feuilles_dpgf=b.get("feuilles_dpgf", []),
            nb_logements_total=b.get("nb_logements_total", 0),
            types_logements=b.get("types_logements", {}),
            systeme_chauffage=b.get("systeme_chauffage", ""),
            notes=b.get("notes", ""),
        )
        for b in config["programme"]
    ]


# ============================================================
# EXTRACTION CCTP (bytes → assertions)
# ============================================================

def extraire_cctp(file_bytes: bytes, config: dict) -> list[AssertionCCTP]:
    doc = Document(io.BytesIO(file_bytes))
    assertions = []

    section_to_bat = {}
    for cctp_section, dpgf_feuilles in config.get("mapping_batiments", {}).items():
        num = cctp_section.replace("CCTP_section_", "")
        if isinstance(dpgf_feuilles, list):
            bat_label = "_".join(dpgf_feuilles)
        else:
            bat_label = str(dpgf_feuilles)
        section_to_bat[num] = bat_label

    current_section = ""
    current_titre = ""
    current_contenu = []
    capture = False
    SECTIONS_A_IGNORER = {"1", "2"}

    def sauvegarder():
        nonlocal current_section, current_titre, current_contenu
        if not current_section or not current_titre:
            return
        num_principal = current_section.split(".")[0]
        if num_principal in SECTIONS_A_IGNORER:
            return
        bat = section_to_bat.get(num_principal, f"SECTION_{num_principal}")

        type_can = forme_canonique_souple(current_titre)
        assertions.append(AssertionCCTP(
            section=current_section,
            titre_article=current_titre,
            texte=current_titre,
            batiment=bat,
            type_canonique=type_can,
            puissance_kw=extraire_puissance(current_titre),
            diametre_dn=extraire_diametre(current_titre),
        ))

        for para in current_contenu:
            p = para.strip()
            if len(p) < 10:
                continue
            tc = forme_canonique_souple(p)
            if tc is not None:
                assertions.append(AssertionCCTP(
                    section=current_section,
                    titre_article=current_titre,
                    texte=p,
                    batiment=bat,
                    type_canonique=tc,
                    puissance_kw=extraire_puissance(p),
                    diametre_dn=extraire_diametre(p),
                ))

    for para in doc.paragraphs:
        texte = para.text.strip()
        if not texte:
            continue
        style = para.style.name if para.style else ""

        if style == "Heading 1":
            sauvegarder()
            current_contenu = []
            current_section = ""
            current_titre = texte
            capture = not any(x in texte.upper() for x in
                ["GENERALITES", "GÉNÉRALITÉS", "ETUDES D'EXECUTION"])
        elif style in ("Heading 2", "Heading 3") and capture:
            sauvegarder()
            current_contenu = []
            match = re.match(r'^(\d+(?:\.\d+)*)', texte)
            current_section = match.group(1) if match else ""
            current_titre = texte
        elif capture and current_titre:
            current_contenu.append(texte)

    sauvegarder()
    return assertions


# ============================================================
# EXTRACTION DPGF (bytes → lignes)
# ============================================================

def extraire_dpgf(file_bytes: bytes, config: dict) -> list[LigneDPGF]:
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    lignes = []

    feuille_to_bat = {}
    for cctp_section, dpgf_feuilles in config.get("mapping_batiments", {}).items():
        if isinstance(dpgf_feuilles, list):
            for f in dpgf_feuilles:
                feuille_to_bat[f] = cctp_section
        else:
            feuille_to_bat[str(dpgf_feuilles)] = cctp_section

    for nom_feuille in wb.sheetnames:
        if nom_feuille.upper() in ("RECAP", "RÉCAPITULATIF", "TOTAL"):
            continue
        ws = wb[nom_feuille]
        bat = feuille_to_bat.get(nom_feuille, nom_feuille)

        for num_ligne, row in enumerate(ws.iter_rows(values_only=True), 1):
            desig = str(row[0]).strip() if row and row[0] else ""
            if not desig or desig == "None":
                continue
            if est_ligne_exclue(desig):
                continue
            if desig.upper() in ("DÉSIGNATION", "DESIGNATION", "LIBELLÉ"):
                continue

            lignes.append(LigneDPGF(
                designation=desig,
                feuille=nom_feuille,
                ligne=num_ligne,
                batiment=bat,
                type_canonique=forme_canonique_souple(desig),
                puissance_kw=extraire_puissance(desig),
                diametre_dn=extraire_diametre(desig),
            ))

    return lignes


# ============================================================
# SCORING + APPARIEMENT
# ============================================================

def scorer_paire(assertion: AssertionCCTP, ligne: LigneDPGF) -> float:
    score = 0.0
    if assertion.type_canonique and ligne.type_canonique:
        if assertion.type_canonique == ligne.type_canonique:
            score += 0.5
        else:
            return 0.0

    mots_a = set(normaliser(assertion.texte).split())
    mots_l = set(normaliser(ligne.designation).split())
    communs = {m for m in (mots_a & mots_l) if len(m) > 3}
    total = max(len(mots_a), 1)
    score += min(len(communs) / total * 0.6, 0.3)

    if assertion.puissance_kw and ligne.puissance_kw:
        if assertion.puissance_kw == ligne.puissance_kw:
            score += 0.1
    if assertion.diametre_dn and ligne.diametre_dn:
        if assertion.diametre_dn == ligne.diametre_dn:
            score += 0.1
    return score


def apparier(assertions, lignes_dpgf, config):
    paires = []
    dpgf_par_section = {}
    for ligne in lignes_dpgf:
        dpgf_par_section.setdefault(ligne.batiment, []).append(ligne)

    for assertion in assertions:
        num_p = assertion.section.split(".")[0] if assertion.section else ""
        section_key = f"CCTP_section_{num_p}"

        feuilles = config.get("mapping_batiments", {}).get(section_key, [])
        if isinstance(feuilles, str):
            feuilles = [feuilles]

        candidats = []
        for f in feuilles:
            candidats.extend(dpgf_par_section.get(f, []))
        candidats.extend(dpgf_par_section.get(section_key, []))

        associees = []
        if assertion.type_canonique:
            for ligne in candidats:
                if ligne.type_canonique == assertion.type_canonique:
                    associees.append(ligne)

        if not associees:
            scored = [(l, scorer_paire(assertion, l)) for l in candidats]
            scored = [(l, s) for l, s in scored if s >= 0.4]
            scored.sort(key=lambda x: x[1], reverse=True)
            associees = [l for l, _ in scored[:5]]

        paires.append((assertion, associees))
    return paires


# ============================================================
# MOTEUR D'ALERTES
# ============================================================

def detecter_alertes(assertions: list[AssertionCCTP],
                     lignes_dpgf: list[LigneDPGF],
                     config: dict,
                     utiliser_ia: bool = False,
                     programme: list[ProgrammeBatiment] = None) -> list[Alerte]:
    alertes = []
    signatures_vues = set()

    def ajouter(a: Alerte):
        sig = a.signature
        if sig not in signatures_vues:
            signatures_vues.add(sig)
            alertes.append(a)

    # ─── PASSE 0 : R5 — désignations incertaines ───
    for ligne in lignes_dpgf:
        motif_r5 = est_designation_incertaine(ligne.designation)
        if motif_r5:
            ajouter(Alerte(
                code="INCERTAIN", criticite="INCERTAIN",
                batiment=ligne.feuille,
                dpgf_feuille=ligne.feuille, dpgf_ligne=ligne.ligne,
                dpgf_texte=ligne.designation,
                motif=motif_r5, regle="R5", methode="Python", confiance=90,
            ))

    # ─── PASSE 1 : C01 — assertion CCTP sans correspondance DPGF ───
    dpgf_types = {}
    for ligne in lignes_dpgf:
        dpgf_types.setdefault(ligne.batiment, set())
        if ligne.type_canonique:
            dpgf_types[ligne.batiment].add(ligne.type_canonique)

    for assertion in assertions:
        if not assertion.type_canonique:
            continue
        num_p = assertion.section.split(".")[0] if assertion.section else ""
        section_key = f"CCTP_section_{num_p}"

        feuilles = config.get("mapping_batiments", {}).get(section_key, [])
        if isinstance(feuilles, str):
            feuilles = [feuilles]

        trouve = False
        for f in feuilles:
            if assertion.type_canonique in dpgf_types.get(f, set()):
                trouve = True
                break
            if assertion.type_canonique in dpgf_types.get(section_key, set()):
                trouve = True
                break

        if not trouve:
            ajouter(Alerte(
                code="C01", criticite="MAJEUR",
                batiment=assertion.batiment,
                cctp_section=assertion.section,
                cctp_texte=assertion.texte[:200],
                motif=f"Type '{assertion.type_canonique}' absent du DPGF",
                regle="Couverture bidirectionnelle",
                methode="Python", confiance=85,
            ))

    # ─── PASSE 2 : C02 — ligne DPGF orpheline ───
    cctp_types = {}
    for a in assertions:
        cctp_types.setdefault(a.batiment, set())
        if a.type_canonique:
            cctp_types[a.batiment].add(a.type_canonique)

    for ligne in lignes_dpgf:
        if not ligne.type_canonique:
            continue
        section_key = ligne.batiment
        feuilles = config.get("mapping_batiments", {}).get(section_key, [])
        if isinstance(feuilles, str):
            feuilles = [feuilles]

        trouve = False
        if isinstance(feuilles, list) and feuilles:
            bat_label = "_".join(feuilles)
            if ligne.type_canonique in cctp_types.get(bat_label, set()):
                trouve = True

        if not trouve:
            ajouter(Alerte(
                code="C02", criticite="MINEUR",
                batiment=ligne.feuille,
                dpgf_feuille=ligne.feuille, dpgf_ligne=ligne.ligne,
                dpgf_texte=ligne.designation[:200],
                motif=f"Type '{ligne.type_canonique}' absent du CCTP",
                regle="Couverture bidirectionnelle",
                methode="Python", confiance=70,
            ))

    # ─── PASSE 3 : C03 / C04 / C05 — comparaison par paires ───
    paires = apparier(assertions, lignes_dpgf, config)

    for assertion, lignes_assoc in paires:
        for ligne in lignes_assoc:
            if sont_equivalents(assertion.texte, ligne.designation):
                continue

            motif_ne = sont_non_equivalents(assertion.texte, ligne.designation)
            if motif_ne:
                ajouter(Alerte(
                    code="C03", criticite="CRITIQUE",
                    batiment=assertion.batiment,
                    cctp_section=assertion.section,
                    cctp_texte=assertion.texte[:200],
                    dpgf_feuille=ligne.feuille, dpgf_ligne=ligne.ligne,
                    dpgf_texte=ligne.designation[:200],
                    motif=motif_ne,
                    regle="R1 — Changement technologie",
                    methode="Python", confiance=95,
                ))
                continue

            res_p = comparer_puissances(assertion.texte, ligne.designation)
            if res_p and res_p.get("alerte"):
                ajouter(Alerte(
                    code="C05", criticite=res_p.get("criticite", "MAJEUR"),
                    batiment=assertion.batiment,
                    cctp_section=assertion.section,
                    cctp_texte=assertion.texte[:200],
                    dpgf_feuille=ligne.feuille, dpgf_ligne=ligne.ligne,
                    dpgf_texte=ligne.designation[:200],
                    motif=res_p["detail"],
                    regle="R4 — Écart puissance",
                    methode="Python", confiance=90,
                ))
                continue

            res_m = comparer_marques(assertion.texte, ligne.designation)
            if res_m["alerte"] == "C04":
                ajouter(Alerte(
                    code="C04", criticite="MAJEUR",
                    batiment=assertion.batiment,
                    cctp_section=assertion.section,
                    cctp_texte=assertion.texte[:200],
                    dpgf_feuille=ligne.feuille, dpgf_ligne=ligne.ligne,
                    dpgf_texte=ligne.designation[:200],
                    motif=res_m["detail"],
                    regle="T3 — Marque différente",
                    methode="Python", confiance=85,
                ))
                continue

    return alertes
